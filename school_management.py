# High Tech School Management System
# A desktop application to manage student and teacher records, payments, and activity logs.
# Uses CustomTkinter for the UI and OpenPyXL for Excel-based data storage.

import customtkinter as ctk
import tkinter as tk
from tkinter import messagebox
import openpyxl
import os
import re
import json
import logging
from datetime import datetime
import sys

# ------------------- Setup and Configuration -------------------

# Configure logging to track errors in a file
logging.basicConfig(filename="error_log.txt", level=logging.DEBUG, format="%(asctime)s - %(levelname)s - %(message)s")

# Set CustomTkinter appearance
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

# Constants for the application
SCHOOL_NAME = "High Tech School"
XLSX_FILE = "school_data.xlsx"
SETTINGS_FILE = "settings.json"
COLORS = {
    "bg": "#1e1e2e",      # Background color
    "frame": "#3b3b4f",   # Frame color
    "accent": "#3b82f6",  # Button color
    "hover": "#60a5fa",   # Hover color
    "text": "#d1d5db",    # Text color
    "paid": "#22c55e",    # Paid status color
    "pending": "#ef4444"  # Pending status color
}
MONTHS = ["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]
YEARS = [str(y) for y in range(2025, 2036)]

# Initialize the main window
try:
    root = ctk.CTk()
    root.title("High Tech School Management System")
    root.geometry("1000x600")
    root.minsize(800, 600)
except Exception as e:
    logging.error(f"Failed to initialize main window: {e}")
    messagebox.showerror("Error", "Failed to start the application")
    sys.exit(1)

# ------------------- Excel Utilities -------------------

def load_workbook():
    """Load the Excel workbook with retry logic."""
    for _ in range(3):
        try:
            return openpyxl.load_workbook(XLSX_FILE)
        except Exception as e:
            logging.error(f"Failed to load workbook: {e}")
    messagebox.showerror("Error", "Failed to load Excel file after multiple attempts")
    return None

def save_workbook(workbook):
    """Save the Excel workbook with retry logic."""
    for _ in range(3):
        try:
            workbook.save(XLSX_FILE)
            return True
        except Exception as e:
            logging.error(f"Failed to save workbook: {e}")
    messagebox.showerror("Error", "Failed to save Excel file after multiple attempts")
    return False

# ------------------- Settings Management -------------------

def load_settings():
    """Load settings from JSON file or return defaults."""
    defaults = {
        "student_id_prefix": "STU-",
        "teacher_id_prefix": "TCH-",
        "student_custom_fields": [],
        "teacher_custom_fields": []
    }
    if os.path.exists(SETTINGS_FILE):
        try:
            with open(SETTINGS_FILE, "r") as file:
                return json.load(file)
        except Exception as e:
            logging.error(f"Failed to load settings: {e}")
            return defaults
    return defaults

def save_settings(settings_data):
    """Save settings to JSON file."""
    try:
        with open(SETTINGS_FILE, "w") as file:
            json.dump(settings_data, file, indent=4)
    except Exception as e:
        logging.error(f"Failed to save settings: {e}")
        messagebox.showerror("Error", f"Failed to save settings: {e}")

settings = load_settings()

# ------------------- Excel Initialization -------------------

def initialize_excel():
    """Set up the Excel file with required sheets and headers."""
    workbook = openpyxl.Workbook() if not os.path.exists(XLSX_FILE) else load_workbook()
    if not workbook:
        sys.exit(1)
    
    sheets = {
        "Students": ["ID", "Name", "Class", "Section", "Year", "Primary Contact", "Secondary Contact", "Tuition_Amount"] + settings["student_custom_fields"],
        "Teachers": ["ID", "Name", "Session_Year", "Primary Contact", "Secondary Contact", "Salary_Amount"] + settings["teacher_custom_fields"],
        "Student_Payments": ["Student_ID", "Amount", "Status", "Month", "Year"],
        "Teacher_Payments": ["Teacher_ID", "Amount", "Status", "Month", "Year"],
        "Activity_Log": ["Timestamp", "Action", "Person_ID", "Details"]
    }
    
    for sheet_name, headers in sheets.items():
        if sheet_name not in workbook.sheetnames:
            workbook.create_sheet(sheet_name).append(headers)
        else:
            sheet = workbook[sheet_name]
            if [cell.value for cell in sheet[1]] != headers:
                workbook.remove(sheet)
                workbook.create_sheet(sheet_name).append(headers)
    
    save_workbook(workbook)

initialize_excel()

# ------------------- Activity Logging -------------------

def log_activity(action, person_id, details):
    """Log an action to the Activity_Log sheet."""
    workbook = load_workbook()
    if not workbook:
        return
    sheet = workbook["Activity_Log"]
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    sheet.append([timestamp, action, person_id, details])
    save_workbook(workbook)

# ------------------- Data Models -------------------

class Payment:
    """Represents a payment record for a student or teacher."""
    def __init__(self, person_id, amount, status="Pending", month=datetime.now().strftime("%B"), year=str(datetime.now().year)):
        self.person_id = person_id
        self.amount = float(amount)
        self.status = status
        self.month = month
        self.year = year

    def to_list(self):
        return [self.person_id, self.amount, self.status, self.month, self.year]

    @staticmethod
    def load_payments(sheet_name):
        """Load payment records from the specified sheet."""
        workbook = load_workbook()
        if not workbook:
            return []
        sheet = workbook[sheet_name]
        payments = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row and len(row) >= 5 and all(x is not None for x in row[:5]):
                payments.append(Payment(*row[:5]))
            else:
                logging.debug(f"Invalid payment row in {sheet_name}: {row}")
        return payments

    @staticmethod
    def save_payments(sheet_name, payments):
        """Save payment records to the specified sheet."""
        workbook = load_workbook()
        if not workbook:
            return False
        if sheet_name not in workbook.sheetnames:
            workbook.create_sheet(sheet_name).append(
                ["Student_ID" if sheet_name == "Student_Payments" else "Teacher_ID", "Amount", "Status", "Month", "Year"]
            )
        sheet = workbook[sheet_name]
        sheet.delete_rows(2, sheet.max_row)
        for payment in payments:
            sheet.append(payment.to_list())
        return save_workbook(workbook)

class Person:
    """Base class for students and teachers."""
    def __init__(self, id, name, primary_contact, secondary_contact="", custom_fields=None):
        self.id = id
        self.name = name
        self.primary_contact = primary_contact
        self.secondary_contact = secondary_contact
        self.custom_fields = custom_fields or []

    def to_list(self):
        return [self.id, self.name, self.primary_contact, self.secondary_contact] + self.custom_fields

class Student(Person):
    """Represents a student with class, section, and tuition details."""
    def __init__(self, id, name, class_name, section, year, primary_contact, secondary_contact, tuition_amount, custom_fields=None):
        super().__init__(id, name, primary_contact, secondary_contact, custom_fields)
        self.class_name = class_name
        self.section = section
        self.year = year
        self.tuition_amount = float(tuition_amount)

    def to_list(self):
        return [self.id, self.name, self.class_name, self.section, self.year, self.primary_contact, self.secondary_contact, self.tuition_amount] + self.custom_fields

class Teacher(Person):
    """Represents a teacher with session year and salary details."""
    def __init__(self, id, name, session_year, primary_contact, secondary_contact, salary_amount, custom_fields=None):
        super().__init__(id, name, primary_contact, secondary_contact, custom_fields)
        self.session_year = session_year
        self.salary_amount = float(salary_amount)

    def to_list(self):
        return [self.id, self.name, self.session_year, self.primary_contact, self.secondary_contact, self.salary_amount] + self.custom_fields

# ------------------- Data Operations -------------------

def load_person_data(sheet_name, is_student=True):
    """Load student or teacher data from the Excel sheet."""
    workbook = load_workbook()
    if not workbook:
        return []
    sheet = workbook[sheet_name]
    person_data = []
    for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if row and len(row) >= (8 if is_student else 6) and all(row[i] is not None for i in range(7 if is_student else 5)):
            try:
                if is_student and not re.match(r"^\d{4}$", str(row[4])):
                    logging.debug(f"Skipping row {row_num} in {sheet_name}: invalid year {row[4]}")
                    continue
                secondary_contact = row[6] if is_student else row[4]
                secondary_contact = secondary_contact if secondary_contact is not None else ""
                args = row[:(8 if is_student else 6)]
                args = args[:6] + (secondary_contact,) + args[7:8] if is_student else args[:4] + (secondary_contact,) + args[5:6]
                person_data.append((Student if is_student else Teacher)(*args, row[8:] if is_student else row[6:]))
            except ValueError as e:
                logging.debug(f"Skipping row {row_num} in {sheet_name}: {e}")
        else:
            logging.debug(f"Invalid row in {sheet_name}: {row}")
    return person_data

def save_person_data(sheet_name, person_data):
    """Save student or teacher data to the Excel sheet."""
    workbook = load_workbook()
    if not workbook:
        return False
    sheet = workbook[sheet_name]
    sheet.delete_rows(2, sheet.max_row)
    for person in person_data:
        sheet.append(person.to_list())
    return save_workbook(workbook)

def generate_unique_id(prefix, existing_ids):
    """Generate a unique ID with the given prefix."""
    max_num = max((int(id[len(prefix):]) for id in existing_ids if id.startswith(prefix) and id[len(prefix):].isdigit()), default=0)
    return f"{prefix}{max_num + 1:03d}"

def validate_input_data(input_data, is_student=True, custom_fields=None, current_id=None):
    """Validate input data for a student or teacher."""
    custom_fields = custom_fields or []
    patterns = [
        (r"^[A-Za-z0-9-]+$", "Invalid ID (use alphanumeric and hyphens)"),
        (r"^[A-Za-z\s]+$", "Invalid Name (use letters only)"),
        (r"^[0-9A-Za-z]+$", "Invalid Class" if is_student else "Invalid Session Year"),
        (r"^[A-Za-z]+$", "Invalid Section" if is_student else "Invalid Primary Contact"),
        (r"^\d{4}$", "Invalid Year (use 4-digit year)" if is_student else "Invalid Primary Contact"),
        (r"^\+?\d{10,15}$", "Invalid Primary Contact (use phone number, e.g., +1234567890)")
    ]
    num_checks = 6 if is_student else 5
    for i, (pattern, error) in enumerate(patterns[:num_checks]):
        if i < len(input_data) and (not input_data[i] or not re.match(pattern, str(input_data[i]))):
            return False, error
    secondary_contact_idx = 6 if is_student else 4
    if secondary_contact_idx < len(input_data) and input_data[secondary_contact_idx] and not re.match(r"^\+?\d{10,15}$", str(input_data[secondary_contact_idx])):
        return False, "Invalid Secondary Contact (use phone number, e.g., +1234567890, or leave empty)"
    amount_idx = 7 if is_student else 5
    try:
        float(input_data[amount_idx])
        if float(input_data[amount_idx]) <= 0:
            return False, f"Invalid {'Tuition' if is_student else 'Salary'} Amount (must be positive)"
    except ValueError:
        return False, f"Invalid {'Tuition' if is_student else 'Salary'} Amount (use numbers)"
    for i, field in enumerate(custom_fields, amount_idx + 1):
        if i < len(input_data) and input_data[i] and not re.match(r"^[A-Za-z0-9\s]+$", str(input_data[i])):
            return False, f"Invalid {field} (use alphanumeric only)"
    return True, ""

# ------------------- UI Components -------------------

def get_class_counts(students):
    """Count students per class for filtering."""
    class_counts = {}
    for student in students:
        class_counts[student.class_name] = class_counts.get(student.class_name, 0) + 1
    return sorted(class_counts.items(), key=lambda x: x[0])

def delete_person(tab_name):
    """Delete a selected student or teacher."""
    config = tab_configs[tab_name]
    selected_id = config.get("selected_id")
    if not selected_id:
        messagebox.showerror("Error", f"Please select a {tab_name[:-1].lower()} to delete")
        return
    selected_idx = next((i for i, p in enumerate(config["person_data"]) if p.id == selected_id), None)
    if selected_idx is None:
        messagebox.showerror("Error", f"Selected {tab_name[:-1].lower()} not found")
        return
    person = config["person_data"][selected_idx]
    if not messagebox.askyesno("Confirm", f"Are you sure you want to delete {person.name} (ID: {person.id})?"):
        return
    config["person_data"].pop(selected_idx)
    config["payments"] = [p for p in config["payments"] if p.person_id != person.id]
    config["selected_id"] = None
    save_person_data(tab_name, config["person_data"])
    Payment.save_payments(f"{tab_name}_Payments", config["payments"])
    log_activity(f"Delete {tab_name[:-1]}", person.id, f"Deleted {person.name}")
    update_person_cards(tab_name)
    update_dashboard()
    messagebox.showinfo("Success", f"{tab_name[:-1]} {person.name} deleted successfully")

def create_form_section(frame, title, fields, entries):
    """Create a section in the form modal with labeled entries."""
    ctk.CTkLabel(frame, text=title, font=("Roboto", 12, "bold")).pack(pady=5)
    for label, value, disabled, placeholder in fields:
        ctk.CTkLabel(frame, text=label, font=("Roboto", 11)).pack(pady=2)
        entry = ctk.CTkEntry(frame, width=200, placeholder_text=placeholder)
        if value:
            entry.insert(0, value)
        if disabled:
            entry.configure(state="disabled")
        entry.pack(pady=2)
        entries.append(entry)

def create_payment_status_section(frame, tab_name, person, config):
    """Create the payment status section in the form modal."""
    ctk.CTkLabel(frame, text="Payment Status", font=("Roboto", 12, "bold")).pack(pady=5)
    ctk.CTkLabel(frame, text="Payment Period", font=("Roboto", 11)).pack(pady=2)
    ctk.CTkOptionMenu(frame, variable=config["month_year_var"], values=[f"{m} {y}" for y in YEARS for m in MONTHS], width=200).pack(pady=2)
    month, year = config["month_year_var"].get().split()
    payment = next((p for p in config["payments"] if p.person_id == person.id and p.month == month and p.year == year), None)
    payment_status_var = ctk.StringVar(value=payment.status if payment else "Pending")
    def toggle_payment_status():
        new_status = "Paid" if payment_status_var.get() == "Pending" else "Pending"
        payment_status_var.set(new_status)
        if payment:
            payment.status = new_status
        else:
            payment = Payment(person.id, person.tuition_amount if tab_name == "Students" else person.salary_amount, new_status, month, year)
            config["payments"].append(payment)
        Payment.save_payments(f"{tab_name}_Payments", config["payments"])
        log_activity(f"Toggle Payment {tab_name[:-1]}", person.id, f"Status changed to {new_status} for {month} {year}")
        update_person_cards(tab_name)
    ctk.CTkLabel(frame, text=f"Status for {month} {year}", font=("Roboto", 11)).pack(pady=2)
    ctk.CTkOptionMenu(frame, variable=payment_status_var, values=["Pending", "Paid"], width=100, command=lambda _: toggle_payment_status()).pack(pady=2)

def create_form_modal(tab_name, selected_idx=None):
    """Create a modal for adding or editing a student or teacher."""
    config = tab_configs[tab_name]
    is_student = tab_name == "Students"
    person = config["person_data"][selected_idx] if selected_idx is not None else None
    
    if selected_idx is not None and person is None:
        logging.error(f"No person found for index {selected_idx} in {tab_name}")
        messagebox.showerror("Error", f"No {tab_name[:-1].lower()} selected")
        return

    modal = ctk.CTkToplevel(root)
    modal.title(f"Edit {tab_name[:-1]}: {person.name} (ID: {person.id})" if person else f"Add New {tab_name[:-1]}")
    modal_width = min(350, int(root.winfo_width() * 0.9))
    modal_height = min(500, int(root.winfo_height() * 0.9))
    modal.geometry(f"{modal_width}x{modal_height}")
    modal.transient(root)
    modal.grab_set()

    main_frame = ctk.CTkFrame(modal, fg_color=COLORS["bg"])
    main_frame.pack(fill="both", expand=True)
    main_frame.grid_rowconfigure(0, weight=1)
    main_frame.grid_columnconfigure(0, weight=1)

    scroll_frame = ctk.CTkScrollableFrame(main_frame, fg_color=COLORS["frame"], width=modal_width-20)
    scroll_frame.grid(row=0, column=0, padx=10, pady=10, sticky="nsew")
    
    footer_frame = ctk.CTkFrame(main_frame, fg_color=COLORS["bg"])
    footer_frame.grid(row=1, column=0, sticky="ew", padx=10, pady=5)

    ctk.CTkLabel(scroll_frame, text=f"Editing {person.name} (ID: {person.id})" if person else f"Add New {tab_name[:-1]}", font=("Roboto", 14)).pack(pady=10)
    
    entries = []
    personal_fields = [
        ("ID", person.id if person else generate_unique_id(settings[config["prefix"]], [p.id for p in config["person_data"]]), True, ""),
        ("Name", person.name if person else "", False, ""),
        ("Class" if is_student else "Session Year", person.class_name if is_student and person else (person.session_year if person else ""), False, ""),
        ("Section" if is_student else "Primary Contact", person.section if is_student and person else (person.primary_contact if person else ""), False, "Enter phone number, e.g., +1234567890" if not is_student else ""),
        ("Year" if is_student else "Secondary Contact", person.year if is_student and person else (person.secondary_contact if person else ""), False, "Enter phone number, e.g., +1234567890 (optional)" if not is_student else ""),
        ("Primary Contact" if is_student else None, person.primary_contact if is_student and person else None, False, "Enter phone number, e.g., +1234567890" if is_student else ""),
        ("Secondary Contact" if is_student else None, person.secondary_contact if is_student and person else None, False, "Enter phone number, e.g., +1234567890 (optional)" if is_student else "")
    ]
    personal_fields = [f for f in personal_fields if f[0] is not None]
    create_form_section(scroll_frame, "Personal Info", personal_fields, entries)
    
    financial_fields = [("Tuition_Amount" if is_student else "Salary_Amount", person.tuition_amount if is_student and person else (person.salary_amount if person else ""), False, "")]
    create_form_section(scroll_frame, "Financial Info", financial_fields, entries)

    if settings[config["custom_fields"]]:
        custom_fields = [(field, person.custom_fields[i] if person and len(person.custom_fields) > i else "", False, "") for i, field in enumerate(settings[config["custom_fields"]])]
        create_form_section(scroll_frame, "Custom Fields", custom_fields, entries)

    if person:
        create_payment_status_section(scroll_frame, tab_name, person, config)

    def save():
        input_data = [e.get() for e in entries]
        is_valid, error = validate_input_data(input_data, is_student, settings[config["custom_fields"]], person.id if person else None)
        if not is_valid:
            messagebox.showerror("Error", error)
            return
        existing_ids = [p.id for p in config["person_data"] if not person or p.id != person.id]
        if input_data[0] in existing_ids:
            messagebox.showerror("Error", f"ID {input_data[0]} already exists")
            return
        new_person = config["class"](*input_data)
        action = "Edit" if person else "Add"
        if person and not messagebox.askyesno("Confirm", f"Save changes to {person.name} (ID: {person.id})?"):
            return
        if selected_idx is not None:
            config["person_data"][selected_idx] = new_person
        else:
            config["person_data"].append(new_person)
            month, year = config["month_year_var"].get().split()
            payment = Payment(new_person.id, new_person.tuition_amount if is_student else new_person.salary_amount, "Pending", month, year)
            config["payments"].append(payment)
            Payment.save_payments(f"{tab_name}_Payments", config["payments"])
        save_person_data(tab_name, config["person_data"])
        log_activity(f"{action} {tab_name[:-1]}", new_person.id, f"{action}ed {new_person.name}")
        config["selected_id"] = None
        update_person_cards(tab_name)
        update_dashboard()
        modal.destroy()
        messagebox.showinfo("Success", f"{tab_name[:-1]} {action.lower()}ed successfully")

    ctk.CTkButton(footer_frame, text="Save", command=save, fg_color=COLORS["accent"], width=100).pack(pady=5)

def toggle_payment(tab_name, person_id, month, year):
    """Toggle the payment status for a person."""
    config = tab_configs[tab_name]
    person = next((p for p in config["person_data"] if p.id == person_id), None)
    if not person:
        logging.error(f"Person not found: {person_id} in {tab_name}")
        messagebox.showerror("Error", f"{tab_name[:-1]} not found")
        return
    payment = next((p for p in config["payments"] if p.person_id == person_id and p.month == month and p.year == year), None)
    if not payment:
        payment = Payment(person_id, person.tuition_amount if tab_name == "Students" else person.salary_amount, "Pending", month, year)
        config["payments"].append(payment)
    payment.status = "Paid" if payment.status == "Pending" else "Pending"
    Payment.save_payments(f"{tab_name}_Payments", config["payments"])
    log_activity(f"Toggle Payment {tab_name[:-1]}", person_id, f"Status changed to {payment.status} for {month} {year}")
    update_person_cards(tab_name)
    update_dashboard()
    messagebox.showinfo("Success", f"Payment status for {person.name} updated to {payment.status}")

def create_person_card(tab_name, person, index, month, year):
    """Create a card widget for a student or teacher."""
    config = tab_configs[tab_name]
    payment = next((p for p in config["payments"] if p.person_id == person.id and p.month == month and p.year == year), None)
    status = payment.status if payment else "Pending"
    card = ctk.CTkFrame(config["card_frame"], fg_color=COLORS["frame"], border_width=2, border_color=COLORS["paid"] if status == "Paid" else COLORS["pending"])
    card.index = index
    card.person_id = person.id
    card.selected = config.get("selected_id") == person.id
    def toggle_select(event):
        config["selected_id"] = person.id
        for child in config["card_frame"].winfo_children():
            child.selected = False
            child_payment = next((p for p in config["payments"] if p.person_id == config["person_data"][child.index].id and p.month == month and p.year == year), None)
            child.configure(fg_color=COLORS["frame"], border_color=COLORS["paid"] if child_payment and child_payment.status == "Paid" else COLORS["pending"])
            for widget in child.winfo_children():
                if isinstance(widget, ctk.CTkLabel) and widget.cget("text").startswith("ID:"):
                    widget.configure(font=("Roboto", 12))
                if isinstance(widget, ctk.CTkLabel) and widget.cget("text") == "Selected":
                    widget.destroy()
        card.selected = True
        card.configure(fg_color=COLORS["accent"], border_color=COLORS["hover"])
        for widget in card.winfo_children():
            if isinstance(widget, ctk.CTkLabel) and widget.cget("text").startswith("ID:"):
                widget.configure(font=("Roboto", 12, "bold"), text_color=COLORS["text"])
        ctk.CTkLabel(card, text="Selected", font=("Roboto", 12, "italic"), text_color=COLORS["text"]).pack(pady=2)
        logging.debug(f"Selected card: {card.person_id}, index: {card.index} in {tab_name}")
    card.bind("<Button-1>", toggle_select)
    ctk.CTkLabel(card, text=f"ID: {person.id}", font=("Roboto", 12, "bold" if card.selected else "normal"), text_color=COLORS["text"]).pack(pady=5)
    ctk.CTkLabel(card, text=person.name, font=("Roboto", 14)).pack(pady=5)
    if tab_name == "Students":
        ctk.CTkLabel(card, text=f"Class: {person.class_name} {person.section}", font=("Roboto", 12)).pack(pady=5)
    else:
        ctk.CTkLabel(card, text=f"Session Year: {person.session_year}", font=("Roboto", 12)).pack(pady=5)
    ctk.CTkLabel(card, text=f"Primary Contact: {person.primary_contact}", font=("Roboto", 12)).pack(pady=5)
    ctk.CTkLabel(card, text=f"Secondary Contact: {person.secondary_contact or 'N/A'}", font=("Roboto", 12)).pack(pady=5)
    ctk.CTkLabel(card, text=f"Status: {status}", font=("Roboto", 12), text_color=COLORS["paid"] if status == "Paid" else COLORS["pending"]).pack(pady=5)
    ctk.CTkButton(card, text="Toggle Payment", command=lambda: toggle_payment(tab_name, person.id, month, year), fg_color=COLORS["accent"]).pack(pady=5)
    if card.selected:
        card.configure(fg_color=COLORS["accent"], border_color=COLORS["hover"])
        ctk.CTkLabel(card, text="Selected", font=("Roboto", 12, "italic"), text_color=COLORS["text"]).pack(pady=2)
    return card

def update_person_cards(tab_name):
    """Update the display of person cards in the UI."""
    config = tab_configs[tab_name]
    for widget in config["card_frame"].winfo_children():
        widget.destroy()
    month, year = config["month_year_var"].get().split()
    person_data = config["person_data"]
    if tab_name == "Students":
        query = config.get("search_query", "").lower()
        selected_class = config.get("selected_class")
        if query:
            person_data = [s for s in person_data if any(query in str(getattr(s, attr)).lower() for attr in ["id", "name", "class_name", "section", "primary_contact", "secondary_contact"])]
        if selected_class and selected_class != "All":
            person_data = [s for s in person_data if s.class_name == selected_class]
    for index, person in enumerate(person_data):
        card = create_person_card(tab_name, person, index, month, year)
        card.pack(pady=10, padx=10, fill="x")
    config["canvas"].configure(scrollregion=config["canvas"].bbox("all"))

# ------------------- Dashboard -------------------

def update_dashboard():
    """Update the dashboard with student and teacher payment statistics."""
    month, year = dashboard_month_year_var.get().split()
    student_counts = {"total": 0, "paid": 0, "pending": 0}
    teacher_counts = {"total": 0, "paid": 0, "pending": 0}
    for student in student_data:
        student_counts["total"] += 1
        payment = next((p for p in student_payments if p.person_id == student.id and p.month == month and p.year == year), None)
        student_counts["paid" if payment and payment.status == "Paid" else "pending"] += 1
    for teacher in teacher_data:
        teacher_counts["total"] += 1
        payment = next((p for p in teacher_payments if p.person_id == teacher.id and p.month == month and p.year == year), None)
        teacher_counts["paid" if payment and payment.status == "Paid" else "pending"] += 1
    
    for widget in dashboard_frame.winfo_children():
        widget.destroy()
    
    ctk.CTkLabel(dashboard_frame, text="Dashboard", font=("Roboto", 16)).pack(pady=10)
    ctk.CTkOptionMenu(dashboard_frame, variable=dashboard_month_year_var, values=[f"{m} {y}" for y in YEARS for m in MONTHS], command=lambda _: update_dashboard()).pack(pady=10)
    frame = ctk.CTkFrame(dashboard_frame, fg_color=COLORS["frame"])
    frame.pack(pady=10, padx=10, fill="both", expand=True)
    ctk.CTkLabel(frame, text=f"Students: {student_counts['total']}", font=("Roboto", 12)).pack(pady=5)
    ctk.CTkLabel(frame, text=f"Paid: {student_counts['paid']}", font=("Roboto", 12), text_color=COLORS["paid"]).pack(pady=5)
    ctk.CTkLabel(frame, text=f"Pending: {student_counts['pending']}", font=("Roboto", 12), text_color=COLORS["pending"]).pack(pady=5)
    ctk.CTkLabel(frame, text=f"Teachers: {teacher_counts['total']}", font=("Roboto", 12)).pack(pady=5)
    ctk.CTkLabel(frame, text=f"Paid: {teacher_counts['paid']}", font=("Roboto", 12), text_color=COLORS["paid"]).pack(pady=5)
    ctk.CTkLabel(frame, text=f"Pending: {teacher_counts['pending']}", font=("Roboto", 12), text_color=COLORS["pending"]).pack(pady=5)

# ------------------- Activity Log -------------------

def update_activity_log():
    """Display the activity log in the UI."""
    workbook = load_workbook()
    if not workbook:
        return
    sheet = workbook["Activity_Log"]
    logs = [row for row in sheet.iter_rows(min_row=2, values_only=True) if row and len(row) >= 4]
    
    for widget in activity_frame.winfo_children():
        widget.destroy()
    
    canvas = ctk.CTkCanvas(activity_frame, bg=COLORS["bg"], highlightthickness=0)
    scrollbar = ctk.CTkScrollbar(activity_frame, orientation="vertical", command=canvas.yview)
    canvas.configure(yscrollcommand=scrollbar.set)
    scrollbar.pack(side="right", fill="y")
    canvas.pack(pady=10, padx=10, fill="both", expand=True)
    log_frame = ctk.CTkFrame(canvas, fg_color=COLORS["bg"])
    canvas.create_window((0, 0), window=log_frame, anchor="nw")
    
    ctk.CTkLabel(log_frame, text="Activity Log", font=("Roboto", 16)).pack(pady=10)
    for log in logs:
        ctk.CTkLabel(log_frame, text=f"{log[0]}: {log[1]} - {log[2]} ({log[3]})", font=("Roboto", 12)).pack(pady=5, anchor="w")
    canvas.configure(scrollregion=canvas.bbox("all"))

# ------------------- UI Setup -------------------

try:
    notebook = ctk.CTkTabview(root, fg_color=COLORS["bg"])
    notebook.grid(row=0, column=0, pady=10, padx=10, sticky="nsew")
    root.grid_rowconfigure(0, weight=1)
    root.grid_columnconfigure(0, weight=1)
except Exception as e:
    logging.error(f"Failed to initialize notebook: {e}")
    messagebox.showerror("Error", f"Failed to initialize UI: {e}")
    sys.exit(1)

# Load initial data
student_data = load_person_data("Students", True)
teacher_data = load_person_data("Teachers", False)
student_payments = Payment.load_payments("Student_Payments")
teacher_payments = Payment.load_payments("Teacher_Payments")

tab_configs = {
    "Students": {
        "person_data": student_data,
        "payments": student_payments,
        "prefix": "student_id_prefix",
        "custom_fields": "student_custom_fields",
        "class": Student,
        "month_year_var": ctk.StringVar(value=f"{datetime.now().strftime('%B')} 2025"),
        "search_query": "",
        "selected_class": None,
        "selected_id": None,
        "card_frame": None,
        "canvas": None
    },
    "Teachers": {
        "person_data": teacher_data,
        "payments": teacher_payments,
        "prefix": "teacher_id_prefix",
        "custom_fields": "teacher_custom_fields",
        "class": Teacher,
        "month_year_var": ctk.StringVar(value=f"{datetime.now().strftime('%B')} 2025"),
        "selected_id": None,
        "card_frame": None,
        "canvas": None
    }
}

# Setup tabs for Students and Teachers
try:
    for tab_name in ["Students", "Teachers"]:
        notebook.add(tab_name)
        frame = ctk.CTkFrame(notebook.tab(tab_name))
        frame.grid(row=0, column=0, sticky="nsew")
        notebook.tab(tab_name).grid_rowconfigure(0, weight=1)
        notebook.tab(tab_name).grid_columnconfigure(0, weight=1)
        
        config = tab_configs[tab_name]
        controls = ctk.CTkFrame(frame, fg_color=COLORS["frame"])
        controls.pack(pady=10, padx=10, fill="x")
        
        if tab_name == "Students":
            search_entry = ctk.CTkEntry(controls, width=200, placeholder_text="Search by Name, ID, Class, Section, or Contact")
            search_entry.pack(side="left", padx=5)
            def update_search(event=None):
                config["search_query"] = search_entry.get()
                update_person_cards("Students")
            search_entry.bind("<KeyRelease>", update_search)
            
            class_counts = get_class_counts(config["person_data"])
            class_options = ["All"] + [f"Class {c} ({n} students)" for c, n in class_counts]
            class_var = ctk.StringVar(value="All")
            def update_class_filter(value):
                config["selected_class"] = None if value == "All" else value.split(" (")[0].replace("Class ", "")
                update_person_cards("Students")
            ctk.CTkOptionMenu(controls, variable=class_var, values=class_options, command=update_class_filter).pack(side="left", padx=5)
            ctk.CTkButton(controls, text="Show All", command=lambda: [class_var.set("All"), update_class_filter("All")], fg_color=COLORS["accent"]).pack(side="left", padx=5)
        
        ctk.CTkButton(controls, text=f"Add {tab_name[:-1]}", command=lambda t=tab_name: create_form_modal(t), fg_color=COLORS["accent"]).pack(side="left", padx=5)
        def edit_command(t=tab_name):
            config = tab_configs[t]
            selected_id = config.get("selected_id")
            if not selected_id:
                messagebox.showerror("Error", f"Please select a {t[:-1].lower()} to edit")
                logging.debug(f"No {t[:-1].lower()} selected for edit, selected_id: {selected_id}")
                return
            selected_idx = next((i for i, p in enumerate(config["person_data"]) if p.id == selected_id), None)
            if selected_idx is None:
                messagebox.showerror("Error", f"Selected {t[:-1].lower()} not found")
                logging.debug(f"Selected {t[:-1].lower()} not found, selected_id: {selected_id}")
                return
            logging.debug(f"Editing {t[:-1].lower()} ID: {selected_id}, index: {selected_idx}")
            create_form_modal(t, selected_idx)
        ctk.CTkButton(controls, text=f"Edit {tab_name[:-1]}", command=edit_command, fg_color=COLORS["accent"]).pack(side="left", padx=5)
        ctk.CTkButton(controls, text=f"Delete {tab_name[:-1]}", command=lambda t=tab_name: delete_person(t), fg_color=COLORS["accent"]).pack(side="left", padx=5)
        ctk.CTkOptionMenu(controls, variable=config["month_year_var"], values=[f"{m} {y}" for y in YEARS for m in MONTHS], command=lambda _: update_person_cards(tab_name)).pack(side="left", padx=5)
        
        canvas = ctk.CTkCanvas(frame, bg=COLORS["bg"], highlightthickness=0)
        scrollbar = ctk.CTkScrollbar(frame, orientation="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        canvas.pack(pady=10, padx=10, fill="both", expand=True)
        config["card_frame"] = ctk.CTkFrame(canvas, fg_color=COLORS["bg"])
        canvas.create_window((0, 0), window=config["card_frame"], anchor="nw")
        config["canvas"] = canvas
        update_person_cards(tab_name)
except Exception as e:
    logging.error(f"Failed to setup tabs: {e}")
    messagebox.showerror("Error", f"Failed to setup tabs: {e}")
    sys.exit(1)

# Setup Dashboard tab
try:
    notebook.add("Dashboard")
    dashboard_frame = ctk.CTkFrame(notebook.tab("Dashboard"))
    dashboard_frame.grid(row=0, column=0, sticky="nsew")
    dashboard_month_year_var = ctk.StringVar(value=f"{datetime.now().strftime('%B')} 2025")
    update_dashboard()
except Exception as e:
    logging.error(f"Failed to setup Dashboard tab: {e}")
    messagebox.showerror("Error", f"Failed to setup Dashboard tab: {e}")

# Setup Activity Log tab
try:
    notebook.add("Activity Log")
    activity_frame = ctk.CTkFrame(notebook.tab("Activity Log"))
    activity_frame.grid(row=0, column=0, sticky="nsew")
    update_activity_log()
except Exception as e:
    logging.error(f"Failed to setup Activity Log tab: {e}")
    messagebox.showerror("Error", f"Failed to setup Activity Log tab: {e}")

# Setup Settings tab
def update_settings():
    """Update application settings and refresh data."""
    settings["student_id_prefix"] = student_prefix_entry.get()
    settings["teacher_id_prefix"] = teacher_prefix_entry.get()
    field_name = field_entry.get().strip()
    field_type = field_type_var.get()
    if field_name and field_name not in settings[f"{field_type.lower()}_custom_fields"]:
        settings[f"{field_type.lower()}_custom_fields"].append(field_name)
        initialize_excel()
        tab_configs[field_type]["person_data"].clear()
        tab_configs[field_type]["person_data"].extend(load_person_data(field_type, field_type == "Students"))
        update_person_cards(field_type)
        log_activity("Add Custom Field", field_type, f"Added field: {field_name}")
    save_settings(settings)
    messagebox.showinfo("Success", "Settings updated")

try:
    notebook.add("Settings")
    settings_frame = ctk.CTkFrame(notebook.tab("Settings"))
    settings_frame.grid(row=0, column=0, sticky="nsew")
    ctk.CTkLabel(settings_frame, text="Settings", font=("Roboto", 16)).pack(pady=10)
    student_prefix_entry = ctk.CTkEntry(settings_frame, width=200, placeholder_text="Student ID Prefix (e.g., STU-)")
    student_prefix_entry.insert(0, settings["student_id_prefix"])
    student_prefix_entry.pack(pady=5)
    teacher_prefix_entry = ctk.CTkEntry(settings_frame, width=200, placeholder_text="Teacher ID Prefix (e.g., TCH-)")
    teacher_prefix_entry.insert(0, settings["teacher_id_prefix"])
    teacher_prefix_entry.pack(pady=5)
    field_entry = ctk.CTkEntry(settings_frame, width=200, placeholder_text="New Custom Field Name")
    field_entry.pack(pady=5)
    field_type_var = ctk.StringVar(value="Students")
    ctk.CTkOptionMenu(settings_frame, variable=field_type_var, values=["Students", "Teachers"]).pack(pady=5)
    ctk.CTkButton(settings_frame, text="Save Settings", command=update_settings, fg_color=COLORS["accent"]).pack(pady=10)
except Exception as e:
    logging.error(f"Failed to setup Settings tab: {e}")
    messagebox.showerror("Error", f"Failed to setup Settings tab: {e}")

# Start the application
try:
    root.mainloop()
except Exception as e:
    logging.error(f"Failed to start main loop: {e}")
    messagebox.showerror("Error", f"Failed to start application: {e}")