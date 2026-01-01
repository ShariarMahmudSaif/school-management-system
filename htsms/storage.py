from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook

from .constants import (
    ACTIVITY_SHEET,
    DATA_XLSX_PATH,
    STUDENTS_SHEET,
    STUDENT_PAYMENTS_SHEET,
    TEACHERS_SHEET,
    TEACHER_PAYMENTS_SHEET,
)
from .logger import AppEvent


def _iso_now() -> str:
    return datetime.now().isoformat(timespec="seconds")


def _ensure_sheet_headers(ws, headers: list[str]) -> None:
    # If the sheet is empty (or has an empty first row), ensure row 1 contains headers.
    # A common failure mode is: row 1 is blank, headers were appended to row 2; the app
    # then reads row 1 as headers (all None), causing keys to be None and IDs not saved.
    if ws.max_row < 1 or ws.max_column < 1:
        for col, h in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=h)
        return

    row1_vals = [cell.value for cell in ws[1]]
    if ws["A1"].value is None and all(v is None for v in row1_vals):
        # If row 2 already contains the correct headers, drop the blank row 1.
        if ws.max_row >= 2:
            row2_vals = [ws.cell(row=2, column=c).value for c in range(1, len(headers) + 1)]
            if row2_vals == headers:
                ws.delete_rows(1, 1)
                return
        # Otherwise, write headers into row 1.
        for col, h in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=h)
        return

    existing = [cell.value for cell in ws[1]]
    if existing != headers:
        # Best-effort: if sheet already has data but headers differ, do not overwrite.
        # New columns will be appended when needed.
        for h in headers:
            if h not in existing:
                ws.cell(row=1, column=len(existing) + 1, value=h)
                existing.append(h)


def _cleanup_sheet(ws, headers: list[str]) -> None:
    """Remove empty rows and duplicated header rows after the header.

    This repairs workbooks that accidentally had blank row 1 (headers appended to row 2)
    or had headers appended multiple times.
    """

    if ws.max_row <= 1:
        return

    # Delete bottom-up to avoid skipping rows.
    for row in range(ws.max_row, 1, -1):
        vals = [ws.cell(row=row, column=c).value for c in range(1, len(headers) + 1)]
        if vals == headers:
            ws.delete_rows(row, 1)
            continue
        if all(v is None or v == "" for v in vals):
            ws.delete_rows(row, 1)


@dataclass
class PersonRecord:
    person_id: str
    first_name: str
    last_name: str
    primary_contact: str
    secondary_contact: str
    extra: dict[str, str]


@dataclass
class StudentRecord(PersonRecord):
    class_name: str
    section: str


@dataclass
class TeacherRecord(PersonRecord):
    role: str


class ExcelStore:
    def __init__(self, path: Path = DATA_XLSX_PATH):
        self.path = path
        self._wb = None

    def invalidate_cache(self) -> None:
        """Force the next operation to re-load the workbook from disk.

        This is important for "live" UIs that watch the xlsx file for changes.
        """
        self._wb = None

    def ensure_workbook(self, student_custom_fields: list[str], teacher_custom_fields: list[str]) -> None:
        if self.path.exists():
            wb = load_workbook(self.path)
        else:
            wb = Workbook()
            # remove default sheet
            default = wb.active
            wb.remove(default)

        # Students
        s_headers = [
            "student_id",
            "first_name",
            "last_name",
            "age",
            "class",
            "section",
            "primary_contact",
            "secondary_contact",
            "created_at",
            "updated_at",
        ] + student_custom_fields
        if STUDENTS_SHEET not in wb.sheetnames:
            ws = wb.create_sheet(STUDENTS_SHEET)
        else:
            ws = wb[STUDENTS_SHEET]
        _ensure_sheet_headers(ws, s_headers)
        _cleanup_sheet(ws, s_headers)

        # Teachers
        t_headers = [
            "teacher_id",
            "first_name",
            "last_name",
            "role",
            "primary_contact",
            "secondary_contact",
            "created_at",
            "updated_at",
        ] + teacher_custom_fields
        if TEACHERS_SHEET not in wb.sheetnames:
            ws = wb.create_sheet(TEACHERS_SHEET)
        else:
            ws = wb[TEACHERS_SHEET]
        _ensure_sheet_headers(ws, t_headers)
        _cleanup_sheet(ws, t_headers)

        # Payments
        for sheet, id_col in [
            (STUDENT_PAYMENTS_SHEET, "student_id"),
            (TEACHER_PAYMENTS_SHEET, "teacher_id"),
        ]:
            headers = [id_col, "year", "month", "status", "amount", "updated_at"]
            if sheet not in wb.sheetnames:
                ws = wb.create_sheet(sheet)
            else:
                ws = wb[sheet]
            _ensure_sheet_headers(ws, headers)
            _cleanup_sheet(ws, headers)

        # Activity
        if ACTIVITY_SHEET not in wb.sheetnames:
            ws = wb.create_sheet(ACTIVITY_SHEET)
        else:
            ws = wb[ACTIVITY_SHEET]
        a_headers = ["timestamp", "action", "entity_type", "entity_id", "details"]
        _ensure_sheet_headers(ws, a_headers)
        _cleanup_sheet(ws, a_headers)

        self.path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(self.path)
        self._wb = wb

    def _load(self):
        if self._wb is not None:
            return self._wb
        if not self.path.exists():
            raise FileNotFoundError(str(self.path))
        self._wb = load_workbook(self.path)
        return self._wb

    def _save(self, wb) -> None:
        wb.save(self.path)
        self._wb = wb

    @staticmethod
    def _sheet_to_dicts(ws) -> list[dict[str, Any]]:
        headers = [c.value for c in ws[1]]
        rows: list[dict[str, Any]] = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            if all(v is None for v in r):
                continue
            d = {headers[i]: r[i] for i in range(min(len(headers), len(r)))}
            rows.append(d)
        return rows

    @staticmethod
    def _find_row_by_id(ws, id_header: str, person_id: str) -> int | None:
        headers = [c.value for c in ws[1]]
        try:
            id_idx = headers.index(id_header) + 1
        except ValueError:
            return None
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=id_idx).value == person_id:
                return row
        return None

    def list_students(self) -> list[dict[str, Any]]:
        wb = self._load()
        ws = wb[STUDENTS_SHEET]
        return self._sheet_to_dicts(ws)

    def list_teachers(self) -> list[dict[str, Any]]:
        wb = self._load()
        ws = wb[TEACHERS_SHEET]
        return self._sheet_to_dicts(ws)

    def upsert_student(self, data: dict[str, Any]) -> None:
        wb = self._load()
        ws = wb[STUDENTS_SHEET]
        headers = [c.value for c in ws[1]]
        row = self._find_row_by_id(ws, "student_id", str(data.get("student_id", "")))
        now = _iso_now()
        if row is None:
            # insert
            row_values = []
            for h in headers:
                if h == "created_at":
                    row_values.append(now)
                elif h == "updated_at":
                    row_values.append(now)
                else:
                    row_values.append(data.get(h, ""))
            ws.append(row_values)
        else:
            for col, h in enumerate(headers, start=1):
                if h == "created_at":
                    continue
                if h == "updated_at":
                    ws.cell(row=row, column=col, value=now)
                    continue
                if h in data:
                    ws.cell(row=row, column=col, value=data.get(h, ""))
        self._save(wb)

    def upsert_teacher(self, data: dict[str, Any]) -> None:
        wb = self._load()
        ws = wb[TEACHERS_SHEET]
        headers = [c.value for c in ws[1]]
        row = self._find_row_by_id(ws, "teacher_id", str(data.get("teacher_id", "")))
        now = _iso_now()
        if row is None:
            row_values = []
            for h in headers:
                if h == "created_at":
                    row_values.append(now)
                elif h == "updated_at":
                    row_values.append(now)
                else:
                    row_values.append(data.get(h, ""))
            ws.append(row_values)
        else:
            for col, h in enumerate(headers, start=1):
                if h == "created_at":
                    continue
                if h == "updated_at":
                    ws.cell(row=row, column=col, value=now)
                    continue
                if h in data:
                    ws.cell(row=row, column=col, value=data.get(h, ""))
        self._save(wb)

    def delete_student(self, student_id: str) -> bool:
        wb = self._load()
        ws = wb[STUDENTS_SHEET]
        row = self._find_row_by_id(ws, "student_id", student_id)
        if row is None:
            return False
        ws.delete_rows(row, 1)
        self._save(wb)
        return True

    def delete_teacher(self, teacher_id: str) -> bool:
        wb = self._load()
        ws = wb[TEACHERS_SHEET]
        row = self._find_row_by_id(ws, "teacher_id", teacher_id)
        if row is None:
            return False
        ws.delete_rows(row, 1)
        self._save(wb)
        return True

    def get_payment_record(self, entity: str, person_id: str, year: int, month: int) -> dict[str, Any]:
        """Get full payment record including status and amount."""
        wb = self._load()
        sheet = STUDENT_PAYMENTS_SHEET if entity == "student" else TEACHER_PAYMENTS_SHEET
        id_col = "student_id" if entity == "student" else "teacher_id"
        ws = wb[sheet]
        rows = self._sheet_to_dicts(ws)
        for r in rows:
            if str(r.get(id_col, "")) == person_id and int(r.get("year", 0)) == year and int(r.get("month", 0)) == month:
                return r
        return {"status": "Pending", "amount": 0.0}

    def get_payment_status(self, entity: str, person_id: str, year: int, month: int) -> str:
        rec = self.get_payment_record(entity, person_id, year, month)
        return str(rec.get("status", "Pending") or "Pending")

    def set_payment(self, entity: str, person_id: str, year: int, month: int, status: str, amount: float) -> None:
        """Set payment status and amount for a person/month."""
        wb = self._load()
        sheet = STUDENT_PAYMENTS_SHEET if entity == "student" else TEACHER_PAYMENTS_SHEET
        id_col = "student_id" if entity == "student" else "teacher_id"
        ws = wb[sheet]
        headers = [c.value for c in ws[1]]
        # Find existing row
        rows = self._sheet_to_dicts(ws)
        target_row: int | None = None
        for idx, r in enumerate(rows, start=2):
            if str(r.get(id_col, "")) == person_id and int(r.get("year", 0)) == year and int(r.get("month", 0)) == month:
                target_row = idx
                break

        now = _iso_now()
        if target_row is None:
            data = {id_col: person_id, "year": year, "month": month, "status": status, "amount": amount, "updated_at": now}
            ws.append([data.get(h, "") for h in headers])
        else:
            for col, h in enumerate(headers, start=1):
                if h == "updated_at":
                    ws.cell(row=target_row, column=col, value=now)
                elif h == "status":
                    ws.cell(row=target_row, column=col, value=status)
                elif h == "amount":
                    ws.cell(row=target_row, column=col, value=amount)
        self._save(wb)

    def set_payment_status(self, entity: str, person_id: str, year: int, month: int, status: str) -> None:
        """Backward-compatible: set status only (keeps existing amount or defaults to 0)."""
        rec = self.get_payment_record(entity, person_id, year, month)
        amt = float(rec.get("amount", 0) or 0)
        self.set_payment(entity, person_id, year, month, status, amt)

    def list_all_payments(self, entity: str) -> list[dict[str, Any]]:
        """Get all payment records for an entity type."""
        wb = self._load()
        sheet = STUDENT_PAYMENTS_SHEET if entity == "student" else TEACHER_PAYMENTS_SHEET
        ws = wb[sheet]
        return self._sheet_to_dicts(ws)

    def get_pending_months(self, entity: str, person_id: str, up_to_year: int, up_to_month: int, default_amount: float) -> list[dict[str, Any]]:
        """Get all unpaid months for a person up to given year/month, with amounts.

        Returns list of dicts with keys: year, month, amount.
        If a month has no record, it's considered pending with default_amount.
        """
        all_payments = self.list_all_payments(entity)
        id_col = "student_id" if entity == "student" else "teacher_id"

        # Build lookup of existing payment records for this person
        paid_lookup: dict[tuple[int, int], dict[str, Any]] = {}
        for rec in all_payments:
            if str(rec.get(id_col, "")) == person_id:
                try:
                    y = int(rec.get("year", 0))
                    m = int(rec.get("month", 0))
                    paid_lookup[(y, m)] = rec
                except Exception:
                    pass

        # Determine start year/month (we'll check last 24 months max to avoid infinite loops)
        pending: list[dict[str, Any]] = []
        # Go back up to 24 months
        y, m = up_to_year, up_to_month
        for _ in range(24):
            rec = paid_lookup.get((y, m))
            if rec is None:
                # No record = pending with default amount
                pending.append({"year": y, "month": m, "amount": default_amount})
            elif str(rec.get("status", "")).lower() != "paid":
                amt = float(rec.get("amount", 0) or 0)
                if amt <= 0:
                    amt = default_amount
                pending.append({"year": y, "month": m, "amount": amt})
            # Move to previous month
            m -= 1
            if m < 1:
                m = 12
                y -= 1
        pending.reverse()  # oldest first
        return pending

    def get_total_pending(self, entity: str, person_id: str, up_to_year: int, up_to_month: int, default_amount: float) -> float:
        """Calculate total pending amount (sum of all unpaid months)."""
        pending = self.get_pending_months(entity, person_id, up_to_year, up_to_month, default_amount)
        return sum(p["amount"] for p in pending)

    def payment_stats(self, entity: str, year: int, month: int) -> dict[str, int]:
        people = self.list_students() if entity == "student" else self.list_teachers()
        id_key = "student_id" if entity == "student" else "teacher_id"
        paid = 0
        pending = 0
        for p in people:
            pid = str(p.get(id_key, ""))
            if not pid:
                continue
            st = self.get_payment_status(entity, pid, year, month)
            if st.lower() == "paid":
                paid += 1
            else:
                pending += 1
        return {"paid": paid, "pending": pending, "total": paid + pending}

    def add_event(self, event: AppEvent) -> None:
        wb = self._load()
        ws = wb[ACTIVITY_SHEET]
        ws.append([event.timestamp, event.action, event.entity_type, event.entity_id, event.details])
        self._save(wb)

    def list_events(self, limit: int = 500) -> list[dict[str, Any]]:
        wb = self._load()
        ws = wb[ACTIVITY_SHEET]
        rows = self._sheet_to_dicts(ws)
        return rows[-limit:]
